import argparse
import logging
import os
import re
from datetime import datetime
from openpyxl import load_workbook


TITLE_PREFIX_PATTERN = re.compile(r"^maker\s+wise\s+fuel\s+data\s+of\s+", re.IGNORECASE)
KNOWN_VEHICLE_TYPES = ("motor_car", "motor_cab")
def setup_logger(log_path: str):
    logger = logging.getLogger("rename_check")
    if logger.handlers:
        return logger

    logger.setLevel(logging.INFO)

    formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setFormatter(formatter)
    logger.addHandler(fh)

    sh = logging.StreamHandler()
    sh.setFormatter(formatter)
    logger.addHandler(sh)

    return logger



def normalize_text(value: str) -> str:
    text = str(value)
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


def extract_header_text(sheet):
    # Prefer first row, but fall back to a small scan window.
    for cell in sheet[1]:
        if cell.value and str(cell.value).strip():
            return str(cell.value).strip()

    for row in sheet.iter_rows(min_row=1, max_row=5, min_col=1, max_col=5):
        for cell in row:
            if cell.value and str(cell.value).strip():
                return str(cell.value).strip()

    return ""


def extract_rto_from_header(header_text: str) -> str:
    if not header_text:
        return ""

    text = TITLE_PREFIX_PATTERN.sub("", header_text).strip()

    text = text.strip()
    text = re.sub(r"\s*,\s*[^,]+(\([^)]*\))?\s*$", "", text)

    return text.strip()


def split_filename(base_name: str):
    for suffix in KNOWN_VEHICLE_TYPES:
        if base_name.lower().endswith(f"_{suffix}"):
            return base_name[: -(len(suffix) + 1)], suffix

    if "_" not in base_name:
        return base_name, ""

    rto_name, vehicle_type = base_name.rsplit("_", 1)
    return rto_name, vehicle_type


def get_available_path(path: str) -> str:
    if not os.path.exists(path):
        return path

    root, ext = os.path.splitext(path)
    for i in range(1, 1000):
        candidate = f"{root} (fix-{i}){ext}"
        if not os.path.exists(candidate):
            return candidate

    return path


def check_and_fix_file(file_path: str, dry_run: bool):
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    rto_from_name, vehicle_type = split_filename(base_name)
    is_report_table = base_name.strip().lower() == "reporttable"
    if is_report_table and not vehicle_type:
        vehicle_type = "motor_car"

    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb.active
    header_text = extract_header_text(sheet)
    wb.close()

    rto_from_header = extract_rto_from_header(header_text)

    if not rto_from_header:
        return {
            "file": file_path,
            "status": "no_header",
            "detail": "Header not found",
        }

    name_norm = normalize_text(rto_from_name)
    header_norm = normalize_text(rto_from_header)

    if name_norm == header_norm:
        return {
            "file": file_path,
            "status": "ok",
            "detail": f"Header matches ({rto_from_header})",
        }

    date_match = re.search(r"\([^)]*\)", rto_from_name)
    if date_match:
        date_suffix = f" {date_match.group(0)}"
    elif is_report_table:
        today = datetime.now()
        date_suffix = f"( {today.strftime('%d-%b-%Y').upper()} )"
        date_suffix = f" {date_suffix}"
    else:
        date_suffix = ""
    new_base = f"{rto_from_header}{date_suffix}_{vehicle_type}".strip("_")
    new_path = os.path.join(os.path.dirname(file_path), f"{new_base}.xlsx")
    new_path = get_available_path(new_path)

    if not dry_run:
        os.rename(file_path, new_path)

    return {
        "file": file_path,
        "status": "renamed" if not dry_run else "mismatch",
        "detail": f"{os.path.basename(file_path)} -> {os.path.basename(new_path)}",
        "header": rto_from_header,
        "new_path": new_path,
    }


def get_default_folder():
    final_folder = f"{datetime.now().date()}_RTO_Files"
    return os.path.join(os.path.expanduser("~"), "Downloads", final_folder)


def run_rename_check(folder: str, dry_run: bool = False, log_file: str | None = None):
    if not log_file:
        log_file = f"rename_check_{datetime.now().strftime('%Y-%m-%d')}.log"

    logger = setup_logger(log_file)

    if not os.path.isdir(folder):
        logger.error("Folder not found: %s", folder)
        return

    files = [
        os.path.join(folder, f)
        for f in os.listdir(folder)
        if f.lower().endswith(".xlsx")
    ]

    if not files:
        logger.warning("No .xlsx files found in %s", folder)
        return

    counts = {"ok": 0, "mismatch": 0, "renamed": 0, "no_header": 0}

    for file_path in sorted(files):
        result = check_and_fix_file(file_path, dry_run)
        status = result["status"]
        counts[status] = counts.get(status, 0) + 1

        if status == "ok":
            continue
        elif status == "no_header":
            continue
        elif status == "mismatch":
            logger.info("MISMATCH : %s", result["detail"])
        else:
            logger.info("RENAMED  : %s", result["detail"])

    logger.info(
        "Done. ok=%s mismatch=%s renamed=%s no_header=%s",
        counts["ok"],
        counts["mismatch"],
        counts["renamed"],
        counts["no_header"],
    )


def main():
    parser = argparse.ArgumentParser(
        description="Check RTO Excel headers vs filenames and rename mismatches."
    )
    parser.add_argument(
        "--folder",
        default=get_default_folder(),
        help="Folder containing downloaded RTO .xlsx files",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Only report mismatches, do not rename files",
    )
    parser.add_argument(
        "--log-file",
        default=f"rename_check_{datetime.now().strftime('%Y-%m-%d')}.log",
        help="Log file path",
    )
    args = parser.parse_args()

    run_rename_check(args.folder, args.dry_run, args.log_file)


if __name__ == "__main__":
    main()
